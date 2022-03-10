#!/usr/bin/env python
# coding: utf-8

# # Excel 处理
# 
# 安装包：
# 
# ```shell
# pip install openpyxl
# ```
# 
# ```{rubric} 新建工作表
# ```

# In[1]:


from openpyxl import Workbook
wb = Workbook()


# ```{rubric} 创建 worksheet
# ```
# 
# ```
# 这在默认情况下设置为 0。除非您修改它的值，否则您将始终使用此方法获取第一个工作表。
# ```

# In[2]:


ws = wb.active


# 你可以使用 {meth}`~openpyxl.Workbook.create_sheet` 方法创建新的工作表：

# In[3]:


ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
# or
ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
# or
ws3 = wb.create_sheet("Mysheet", -1) # insert at the penultimate position


# In[ ]:




